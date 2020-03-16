#Created soley by Isaac Pamu Joshi

import os
import sys
import openpyxl as xl
from Location import Location
from Cluster import Cluster
import urllib.request
import PIL.Image
import json
import tkinter
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import *
import time

def form_clusters(S,distance):
	coords=set(S) 
	clusters=[]
	ID_no = 65 #interger form of "A"
	while len(coords):
		assingned = False
		locus=coords.pop()
		for cluster in clusters: #check if it fits in an existing cluster
			for x in cluster.loc_collection:
				if locus.find_distance(x) <= distance and not locus.loc_ID == x.loc_ID and assingned == False:
					cluster.add_loc(locus)
					assingned = True
		if assingned == False: #create new cluster if loc doesn't fit in an existing cluster
			cluster = Cluster()
			cluster.ID = chr(ID_no)
			ID_no = ID_no + 1
			cluster.loc_collection = [x for x in coords if (locus.find_distance(x) <= distance and not locus.loc_ID == x.loc_ID)] 
			cluster.add_loc(locus)
			clusters.append(cluster) 
			for x in cluster.loc_collection:
				if x in coords:
					coords.remove(x)
	return clusters

def process_locs_inputs(wb_name):
	locs = []
	input_wb = xl.load_workbook(filename = wb_name)
	input_wb_main_ws = input_wb.active
	state = ["error","error", "NSW","VIC","QLD","SA","WA","TAS","NT"]
	print(input_wb_main_ws.max_row)
	for i in range(2,input_wb_main_ws.max_row + 1):
		loc = Location()
		if input_wb_main_ws.cell(row=i, column=1).value == None:
			loc.loc_ID = i -1
		else:
			loc.loc_ID = str(input_wb_main_ws.cell(row=i, column=1).value)
		loc.address = str(input_wb_main_ws.cell(row=i, column=2).value)
		loc.postcode = str(input_wb_main_ws.cell(row=i, column=3).value)
		#add co ordinates from bing maps if needed
		if input_wb_main_ws.cell(row=i, column=4).value == None or input_wb_main_ws.cell(row=i, column=5).value == None:
			if loc.address == "None":
				continue
				print("please provide address or lat/long")
			loc.get_coordinates()
			input_wb_main_ws.cell(row=i, column=4).value = loc.lat
			input_wb_main_ws.cell(row=i, column=5).value = loc.long
		else:
			loc.lat = float(input_wb_main_ws.cell(row=i, column=4).value)
			loc.long = float(input_wb_main_ws.cell(row=i, column=5).value)
		#add address from bing maps if needed
		if loc.address == "None":
			print(i)
			print(loc.address)
			loc.get_address()
			input_wb_main_ws.cell(row=i, column=2).value = loc.address
		loc.SAM_ID = str(input_wb_main_ws.cell(row=i, column=6).value)
		loc.status = str(input_wb_main_ws.cell(row=i, column=8).value)
		loc.MRQ = str(input_wb_main_ws.cell(row=i, column=7).value)
		if loc.SAM_ID != "None":
			loc.state = state[int(loc.SAM_ID[0])]
		else:
			loc.state = "None"
		loc.ID = i - 1
		dup = False
		for loc_base in locs:
			if loc_base.loc_ID == loc.loc_ID:
				dup = True
				break
		if dup == False:
			locs.append(loc)
	#wb_name[-5]="1"
	#print("input saving as: " + wb_name)
	#input_wb.save(wb_name)
	input_wb.close()
	return locs

def process_locs_status(locs, filename):
	print("this shouldnt be called")
	sys.exit()
	try:
		wb = xl.load_workbook(filename = filename)
	except Exception as e:
		print(e)
		return
	ws = wb["Sheet1"]
	for loc in locs:
		#print(ws.max_row + 1)
		for i in range(2,ws.max_row + 1):
			#print("comparing" + loc.loc_ID + " and " + str(ws.cell(row=i, column=19).value))
			if loc.loc_ID == str(ws.cell(row=i, column=19).value) and str(ws.cell(row=i, column=18).value) != "None":
				if loc.status == "None":
					loc.status = ws.cell(row=i, column=15).value
				loc.MRQ = ws.cell(row=i, column=18).value
				loc.tech = ws.cell(row=i, column=24).value
				if loc.SAM_ID == "None":
					loc.SAM_ID = ws.cell(row=i, column=26).value
				if loc.address == "None":
					loc.address = ws.cell(row=i, column=20).value
				if loc.state == "None":
					loc.state = ws.cell(row=i, column=25).value
				break

def process_locs_output(wb, locs):
	#list locs
	wb.create_sheet("Locations")
	ws = wb["Locations"]
	field_names = ["Temp ID","Loc ID", "Address", "Latitude", "Longitude", "Postcode", "SAM ID", "State", "Tech", "MRQ", "Status"]
	for x in range(11):
		ws.cell(row=1, column=x + 1).value = field_names[x]
	
	i = 2
	for loc in locs:
		ws.cell(row=i, column=2).value = loc.loc_ID
		ws.cell(row=i, column=3).value = loc.address
		ws.cell(row=i, column=4).value = loc.lat
		ws.cell(row=i, column=5).value = loc.long
		ws.cell(row=i, column=6).value = loc.postcode
		ws.cell(row=i, column=7).value = loc.SAM_ID 
		ws.cell(row=i, column=8).value = loc.state
		ws.cell(row=i, column=1).value = loc.ID
		ws.cell(row=i, column=9).value = loc.tech
		ws.cell(row=i, column=10).value = loc.MRQ
		ws.cell(row=i, column=11).value = loc.status
		i = i + 1

def process_outputs(clusters, locs):
	output_wb = xl.Workbook()
	output_wb_main_ws = output_wb.active
	output_wb_main_ws.Name = "Cluster Distances"

	process_locs_output(output_wb, locs)
	
	#set columns and rows
	i = 2
	for cluster in clusters:
		output_wb_main_ws.cell(row=1, column=i).value = cluster.ID
		output_wb_main_ws.cell(row=i, column=1).value = cluster.ID
		i = i + 1

	#fill cells
	i = 2
	for cluster_base in clusters:
		j = 2
		for cluster_target in clusters:
			output_wb_main_ws.cell(row=i, column=j).value = cluster_base.loc_collection[0].find_distance(cluster_target.loc_collection[0])*111
			#print("distance between " + cluster_base.ID + " " + cluster_target.ID + " " + str(output_wb_main_ws.cell(row=i, column=j).value))
			j = j + 1
		i = i + 1

	for cluster in clusters:
		cluster.add_cluster_sheet(output_wb)

	#save output
	output_wb.save(os.getcwd() + "/output/" + "Distances.xlsx")

def get_clusters_image(clusters):
		bingkey = "AsuwbgYYc0ypzGFmASWf0VMPk9jme369wjy9uxZkddJk-7SGyT_txN0bSB88a1AM"
		Url = "https://dev.virtualearth.net/REST/V1/Imagery/Map/road?"#ma="+str(min_y)+","+str(min_x)+","+str(max_y)+","+str(max_x)#+"&ms=600,1000"#&zoomLevel="#+str(self.width_to_zoom_level())
		Url = Url + "ms=900,900"
		for cluster in clusters:
			centre, width, height = cluster.form_rectangle()
			Url = Url + "&pp=" + str(centre[0]) + "," + str(centre[1]) + ";46;" + str(cluster.ID)
			if len(cluster.loc_collection) > 99:
				continue
			for loc in cluster.loc_collection:
				Url = Url + "&pp=" + str(loc.lat) + "," + str(loc.long) + ";1;" + str(loc.ID)
			
		Url = Url + "&dcl=1"
		Url = Url + "&key=" + bingkey
		#print(Url)
		request = urllib.request.Request(Url)
		#time.sleep(2)
		#response = urllib.request.urlopen(request)
		#opened = False
		i = 0
		while True:
			try:
				response = urllib.request.urlopen(request)
			except Exception as e:
				print(e)
				time.sleep(1)
				i = i + 1
				if i > 10:
					response = urllib.request.urlopen(request)
					print("time out clusters")
					break
			else:
				break
				#opened = True
		map_image = PIL.Image.open(response)	
		filename = "all locs"
		map_image.save(os.getcwd() + "/output/" + str(filename) + ".png")

def get_super_cluster_image(clusters):
		bingkey = "AsuwbgYYc0ypzGFmASWf0VMPk9jme369wjy9uxZkddJk-7SGyT_txN0bSB88a1AM"
		Url = "https://dev.virtualearth.net/REST/V1/Imagery/Map/road?"#ma="+str(min_y)+","+str(min_x)+","+str(max_y)+","+str(max_x)#+"&ms=600,1000"#&zoomLevel="#+str(self.width_to_zoom_level())
		Url = Url + "ms=900,900"
		for cluster in clusters:
			centre, width, height = cluster.form_rectangle()
			Url = Url + "&pp=" + str(centre[0]) + "," + str(centre[1]) + ";46;" + str(cluster.ID)
		Url = Url + "&dcl=1"
		Url = Url + "&key=" + bingkey
		#print(Url)
		request = urllib.request.Request(Url)
		#time.sleep(2)
		#response = urllib.request.urlopen(request)
		#opened = False
		i = 0
		while True:
			try:
				response = urllib.request.urlopen(request)
			except Exception as e:
				print(e)
				time.sleep(1)
				i = i + 1
				if i > 10:
					response = urllib.request.urlopen(request)
					print("time out super cluster")
					break
			else:
				break
				#opened = True
		map_image = PIL.Image.open(response)	
		filename = "all clusters"
		map_image.save(os.getcwd() + "/output/" + filename + ".png")

def get_input_filename(message = "Select file"):
	root = tkinter.Tk()
	root.withdraw()
	filename_input = filedialog.askopenfilename(title = message)
	print ("Loading File Path: ", filename_input)
	root.destroy()
	return filename_input

def get_input_value():
	root = tkinter.Tk()
	root.withdraw()
	myText=tkinter.simpledialog.askstring("Eneter clustering distance","Enter distance in km:",parent=root)
	return int(myText)

def remake_locs(clusters, locs):
	ID_no = 1
	temp_locs = []
	for cluster in clusters:
		#print(cluster.ID)
		for loc_base in cluster.loc_collection:
			for loc_target in locs:
				if loc_base.address == loc_target.address:
					temp_loc = loc_target
					temp_loc.ID = ID_no
					temp_locs.append(temp_loc)
			ID_no = ID_no + 1
	return temp_locs

def prepare_folder():
	folder_path = os.getcwd()+"/output"
	if not os.path.exists(folder_path):
		os.mkdir(folder_path)
	for the_file in os.listdir(folder_path):
		file_path = os.path.join(folder_path, the_file)
		try:
			if os.path.isfile(file_path):
				os.unlink(file_path)
		except Exception as e:
			print(e)

def main():
	#Prepare folders
	prepare_folder()
	
	#get input db filename
	filename_input = get_input_filename("Select Input File")
	distance_lat = get_input_value()/111
	
	print("processing inputs")
	locs = process_locs_inputs(filename_input)

	
	print("forming clusters")
	clusters = form_clusters(locs, distance_lat)
	locs = remake_locs(clusters, locs)
	clusters = form_clusters(locs, distance_lat)

	#get images
	print("creating images")
	i = 0
	for cluster in clusters:
		i = i + 1
		cluster.form_end_points()
		cluster.get_cluster_image()
		print("created image for cluster " + str(i))
	get_clusters_image(clusters)
	get_super_cluster_image(clusters)
	
	#make output excel
	process_outputs(clusters, locs)

	#Finished successfully
	print("Finished successfully")

#def import_packages():
#	pip.main(["install", "openpyxl"]) #nameOfPackage])
#	pip.main(["install", "Pillow"]) #nameOfPackage])


main()
