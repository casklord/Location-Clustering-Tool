import openpyxl as xl
import urllib.request
import PIL.Image
import os
import time

class Cluster:
	def __init__(self):
		self.loc_collection = []
		self.end_points = []
		self.ID = None

	def add_loc(self, loc):
		self.loc_collection.append(loc)

	def set_end_points(self, end_points):
		self.end_points = end_points

	def print(self):
		print("cluster ID is: " + str(self.ID))
		for loc in self.loc_collection:
			loc.print()
		
	def form_end_points(self):
		max_y = float("-inf")
		max_x = float("-inf")
		min_y = float("inf")
		min_x = float("inf")
		for loc in self.loc_collection:
			if loc.long > max_y:
				max_y = loc.long 
			if loc.lat > max_x:
				max_x = loc.lat 
			if loc.long < min_y:
				min_y = loc.long 	
			if loc.lat < min_x:
				min_x = loc.lat 
		self.end_points = [max_y + 1, max_x + 1, min_y - 1, min_x - 1]

	def get_cluster_image(self):
		bingkey = "AsuwbgYYc0ypzGFmASWf0VMPk9jme369wjy9uxZkddJk-7SGyT_txN0bSB88a1AM"
		Url = "https://dev.virtualearth.net/REST/v1/Imagery/Map/Road?"#ma="+str(min_y)+","+str(min_x)+","+str(max_y)+","+str(max_x)#+"&ms=600,1000"#&zoomLevel="#+str(self.width_to_zoom_level())
		Url = Url + "ms=600,1000"
		if len(self.loc_collection) > 99:
				print("cannot visualise all locs")
		i = 0
		for loc in self.loc_collection:
			Url = Url + "&pp=" + str(round(loc.lat,3)) + "," + str(round(loc.long,3)) 
			#print(loc.status)
			if loc.status == "ACCEPTED" or loc.status == "ALLOCATED" or loc.status == "SUBMITTED":
				Url = Url + ";80;" + str(loc.ID) #RED
			elif loc.status == "PENDINGINFO":
				Url = Url + ";88;" + str(loc.ID) #ORANGE
			elif  loc.status == "INPRG" or loc.status == "DONE":
				Url = Url + ";79;" + str(loc.ID) #GREEN
			else:
				Url = Url + ";84;" + str(loc.ID) #PURPLE				
			if i == 75:
				break
			i = i + 1
		centre, width, height = self.form_rectangle()
		Url = Url + "&pp=" + str(centre[0]) + "," + str(centre[1]) + ";46;" + str(self.ID)
		Url = Url + "&dcl=1"
		Url = Url + "&key=" + bingkey
		#print(Url)
		request = urllib.request.Request(Url)
		time.sleep(5)
		response = urllib.request.urlopen(request)
		'''i = 0
		while True:
			try:
				response = urllib.request.urlopen(request)
			except Exception as e:
				print(e)
				time.sleep(1)
				i = i + 1
				if i > 4:
					response = urllib.request.urlopen(request)
					print("time out getting cluster image")
					break
			else:
				break'''
		map_image = PIL.Image.open(response)	
		filename = str("cluster " + self.ID + " with " + str(len(self.loc_collection))) + " locations , and centre at lat, long " + str(round(centre[0],3)) + ", " + str(round(centre[1],3))
		print("saving " + filename)
		map_image.save(os.getcwd() + "/output/" + filename + ".png")

	def form_rectangle(self):
		[max_y, max_x, min_y, min_x] = self.end_points
		centre = [round((max_x + min_x)/2,2), round((max_y + min_y)/2,2)]
		height =  round((max_y - min_y),2)
		width = round((max_x - min_x),2)
		return centre, width, height

	def add_cluster_sheet(self, wb):
		wb.create_sheet(self.ID)
		ws = wb[self.ID]
		field_names = ["Temp ID","Loc ID", "Address", "Latitude", "Longitude", "Postcode", "SAM ID", "State", "Tech", "MRQ", "Status"]
		for x in range(11):
			ws.cell(row=1, column=x + 1).value = field_names[x]
		
		i = 2
		for loc in self.loc_collection:
			ws.cell(row=i, column=2).value = loc.loc_ID
			ws.cell(row=i, column=3).value = loc.address
			ws.cell(row=i, column=4).value = loc.lat
			ws.cell(row=i, column=5).value = loc.long
			ws.cell(row=i, column=6).value = loc.postcode
			ws.cell(row=i, column=7).value = loc.SAM_ID 
			ws.cell(row=i, column=8).value = loc.state
			ws.cell(row=i, column=1).value = loc.ID
			ws.cell(row=i, column=9).value = loc.tech
			ws.cell(row=i, column=10).value = loc.MRQ
			ws.cell(row=i, column=11).value = loc.status
			if len(self.loc_collection) != 1:
				ws.cell(row=1, column=i + 10).value = "Distance to " + str(loc.loc_ID)
			i = i + 1
		if len(self.loc_collection) == 1:
			return
		i = 2
		for loc_base in self.loc_collection:
			j = 12
			for loc_target in self.loc_collection:
				ws.cell(row=i, column=j).value = loc_base.find_distance(loc_target)*111
				j = j + 1
			i = i + 1


